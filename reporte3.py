import sqlite3
import pandas as pd
import re
from datetime import datetime
from fpdf import   FPDF
import os
import locale
import unicodedata
import pandas as pd
from dateutil.relativedelta import relativedelta
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportes.config import name_db, empresas, tablas_por_empresa_reporte1, cargar_json, privadas_dict
from reportes.pdf_utils import ReportePDF, ruta_recurso
from reportes.config_pdf import FUENTE_REGULAR, LOGO_FILENAME, TAMANO_SUBTITULO



locale.setlocale(locale.LC_TIME, "Spanish_Mexico") 



class Reporte3:
    def __init__(self):
        self.name_db = name_db
        self.empresas = empresas
        self.tablas_por_empresa = tablas_por_empresa_reporte1
        self.superficies_lotes = cargar_json("superficies_lotes.json")
        self.config_pagos = cargar_json("mapeo_pagos.json")
        self.mapeo = cargar_json("mapeo_transferencias.json")
        self.construir_catalogo_clientes_global()
        # Un solo patrón para todas las funciones
        self.RE_LOTES = re.compile(r"(LOTE|CONSTRUCCION|LOTE PALMARENA|CONSTRUCCION CASA RESID PALMARENA|LOTE LAS HADAS)", re.IGNORECASE)
    
    def _aliases_por_nombre_ui(self, nombre_ui: str) -> set[str]:
        """
        Devuelve TODOS los nombres (aliases) observados para el/los ID(s) del nombre que el usuario eligió en la UI.
        Regresa en UPPER y normalizados (sin tildes/espacios repetidos) para comparar.
        """
        ids = self._resolver_ids_por_nombre_canonico(nombre_ui)
        if not ids:
            return { self._norm_txt(nombre_ui) }
        aliases = self._obtener_aliases_por_ids(ids)  # ya lo tienes
        if not aliases:
            return { self._norm_txt(nombre_ui) }
        return { self._norm_txt(a) for a in aliases }

    def _ids_por_nombre_ui(self, nombre_ui: str) -> set[int]:
        """Saca los ID(s) asociados al nombre canónico que se muestra en UI."""
        return self._resolver_ids_por_nombre_canonico(nombre_ui) or set()

    def _resolver_ids_global_por_texto(self, nombre_ui: str) -> set[int]:
        """
        Busca IDs cuyo CRAZONSOCIAL coincide (normalizado) con el texto dado
        a lo largo de TODAS las empresas (por si el usuario escribe el alias 'B').
        """
        target = self._norm_txt(nombre_ui)
        ids = set()
        for empresa in self.empresas:
            try:
                tablas = self.tablas_por_empresa[empresa]
                documentos = f'"{tablas["documentos"]}"'
                with sqlite3.connect(self.name_db) as conn:
                    df = pd.read_sql_query(f"""
                        SELECT DISTINCT CIDCLIENTEPROVEEDOR, CRAZONSOCIAL
                        FROM {documentos}
                        WHERE CRAZONSOCIAL IS NOT NULL
                    """, conn)
                if df.empty:
                    continue
                df["__N__"] = df["CRAZONSOCIAL"].astype(str).map(self._norm_txt)
                ids |= set(df.loc[df["__N__"] == target, "CIDCLIENTEPROVEEDOR"].dropna().astype(int).tolist())
            except Exception as e:
                print(f"⚠️ _resolver_ids_global_por_texto error en {empresa}: {e}")
                continue
        return ids


    def _aliases_de_ids_union(self, ids: set[int]) -> set[str]:
        """
        Devuelve todos los aliases (CRAZONSOCIAL) observados para un conjunto de IDs (todas las empresas),
        normalizados a UPPER y sin tildes/espacios redundantes.
        """
        if not ids:
            return set()
        aliases = self._obtener_aliases_por_ids(ids)  # ya une de cache + fallback
        return { self._norm_txt(a) for a in aliases }

        
    
    def obtener_lotes_equivalentes(self, cliente, lote_original):
        """
        Retorna solo el lote base y su versión exacta con ' COMPL' al final,
        si ambas existen y son exactamente iguales salvo por ese sufijo.
        """
        lote_original = lote_original.strip()
        lote_upper = lote_original.upper()
        
        # 1. Lote base sin COMPL (si lo tiene)
        if lote_upper.endswith(" COMPL"):
            lote_base = lote_original[:-6].strip()
        else:
            lote_base = lote_original

        # 2. Versiones posibles
        version_base = lote_base
        version_compl = f"{lote_base} COMPL"

        # 3. Lotes válidos del cliente
        lotes_cliente = self.obtener_lotes_por_cliente(cliente)
        lotes_cliente_upper = {l.strip().upper(): l for l in lotes_cliente}

        # 🧪 Debug
        print(f"🎯 Comparando base: '{version_base}' y COMPL: '{version_compl}'")
        print(f"📋 Lotes del cliente ({cliente}):")
        for l in lotes_cliente:
            print(f"   - {l}")

        # 4. Verificar si ambas versiones existen exactamente
        tiene_base = version_base.strip().upper() in lotes_cliente_upper
        tiene_compl = version_compl.strip().upper() in lotes_cliente_upper

        posibles_lotes = []
        if tiene_base and tiene_compl:
            posibles_lotes = [
                lotes_cliente_upper[version_base.upper()],
                lotes_cliente_upper[version_compl.upper()]
            ]
        elif tiene_base:
            posibles_lotes = [lotes_cliente_upper[version_base.upper()]]
        elif tiene_compl:
            posibles_lotes = [lotes_cliente_upper[version_compl.upper()]]
        else:
            posibles_lotes = [lote_original]

        print("✅ Lotes equivalentes seleccionados:", posibles_lotes)
        return posibles_lotes
    
    def filtrar_lotes_exactos(self, df, posibles_lotes):
        posibles_lotes_limpios = [l.strip().upper() for l in posibles_lotes]
        df = df.copy()
        df["CNOMBREPRODUCTO_LIMPIO"] = df["CNOMBREPRODUCTO"].str.strip().str.upper()
        
        # DEBUG
        print("🎯 Lotes en DataFrame normalizados:")
        print(df["CNOMBREPRODUCTO_LIMPIO"].unique())
        print("🔍 Lotes buscados:")
        print(posibles_lotes_limpios)

        df_filtrado = df[df["CNOMBREPRODUCTO_LIMPIO"].isin(posibles_lotes_limpios)]

        # Validar si se coló alguna construcción
        colados = df_filtrado[df_filtrado["CNOMBREPRODUCTO_LIMPIO"].str.contains("CONSTRUCCION")]
        if not colados.empty:
            print("🚨 ¡Se colaron construcciones!")
            print(colados[["CNOMBREPRODUCTO", "FACTURA", "CTOTAL"]])

        return df_filtrado



    def limpiar_nombre_lote(self, texto):
        return re.sub(r'\s+', ' ', texto.strip().upper().replace('\n', '').replace('\r', ''))


    def _normalizar_fecha(self, df, col):
        """Convierte df[col] a DD/MM/YYYY de forma consistente (dayfirst=True)."""
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
            df[col] = df[col].dt.strftime("%d/%m/%Y")
        return df


    def _calcular_saldo_vencido(self, df_programa, df_pagos_lote, fecha_corte=None):
        """
        Calcula el saldo vencido asignando pagos (hasta fecha_corte) a cuotas por orden de vencimiento (FIFO).

        df_programa: DataFrame con columnas ["Fecha de pago", "Pago mensual"] (EstadoCuenta_PeriodoPagos),
                    "Fecha de pago" puede venir como str; aquí se normaliza.
        df_pagos_lote: DataFrame con columnas ["Fecha de pago", "Importe MXN"] (EstadoCuenta_Pagos filtrado por facturas),
                    "Fecha de pago" = fecha del abono; se normaliza y se filtra <= fecha_corte.
        fecha_corte: datetime.date (por defecto: hoy).

        Retorna: float (saldo vencido >= 0)
        """

        if fecha_corte is None:
            fecha_corte = datetime.now().date()

        # Normalizar fechas del PROGRAMA (vencimientos)
        df_prog = df_programa.copy()
        if "Fecha de pago" not in df_prog.columns or "Pago mensual" not in df_prog.columns:
            print("🚨 _calcular_saldo_vencido: df_programa no tiene columnas necesarias.")
            return 0.0

        df_prog["Fecha de pago"] = pd.to_datetime(df_prog["Fecha de pago"], dayfirst=True, errors="coerce")
        df_prog = df_prog.dropna(subset=["Fecha de pago"])
        df_prog["Pago mensual"] = pd.to_numeric(df_prog["Pago mensual"], errors="coerce").fillna(0.0)
        df_prog = df_prog.sort_values("Fecha de pago").reset_index(drop=True)

        # Normalizar fechas y filtrar pagos HASTA fecha_corte (¡pagos posteriores no reducen vencido!)
        df_pay = df_pagos_lote.copy() if df_pagos_lote is not None else pd.DataFrame(columns=["Fecha de pago", "Importe MXN"])
        if not df_pay.empty:
            # Si tu DF de pagos trae "Fecha de pago" como string dd/mm/yyyy, normalizamos igual
            df_pay["Fecha de pago"] = pd.to_datetime(df_pay["Fecha de pago"], dayfirst=True, errors="coerce")
            df_pay = df_pay.dropna(subset=["Fecha de pago"])
            df_pay = df_pay[df_pay["Fecha de pago"].dt.date <= fecha_corte]
            df_pay["Importe MXN"] = pd.to_numeric(df_pay["Importe MXN"], errors="coerce").fillna(0.0)
            pagos_disponibles = float(df_pay["Importe MXN"].sum())
        else:
            pagos_disponibles = 0.0

        print(f"🗓️ Fecha de corte (vencido): {fecha_corte.strftime('%d/%m/%Y')}")
        print(f"💵 Pagos disponibles hasta corte: {pagos_disponibles:,.2f}")

        # Asignación FIFO de pagos a cuotas
        vencido_total = 0.0
        for _, row in df_prog.iterrows():
            vence = row["Fecha de pago"].date()
            cuota = float(row["Pago mensual"])
            if cuota <= 0:
                continue

            aplicado = min(pagos_disponibles, cuota)
            pendiente = cuota - aplicado
            pagos_disponibles -= aplicado

            # Solo cuentan como vencidas las cuotas con fecha < corte
            if vence < fecha_corte:
                vencido_total += pendiente

        vencido_total = max(0.0, float(vencido_total))
        print(f"📌 Saldo vencido calculado (FIFO): {vencido_total:,.2f}")
        return vencido_total




    def generar_estado_cuenta_completo(self, cliente, lote, empresa_objetivo=None, filtros_forma_pago=None,
                                   incluir_programa=False, tasa_interes=None, enganche=None, meses=None,
                                   formula=None, fecha_inicio=None):

        try:
            print(f"📋 Generando estado de cuenta completo para cliente '{cliente}' y lote '{lote}'...")

            posibles_lotes = self.obtener_lotes_equivalentes(cliente, lote)
            print(f"🔗 Lotes equivalentes encontrados: {posibles_lotes}")

            # 1) Generar resumen (guarda EstadoCuenta_Resumen con columna Empresa)
            resumen_generado = self.generar_tabla_estado_cuenta_cliente(cliente, posibles_lotes)
            if not resumen_generado:
                print("❌ No se pudo generar el estado de cuenta. No se encontró información.")
                return

            # 2) Leer resumen y FILTRAR POR EMPRESA antes de filtrar por lote
            with sqlite3.connect(self.name_db) as conn:
                df_resumen = pd.read_sql_query("SELECT * FROM EstadoCuenta_Resumen", conn)

            # 🔒 Filtro por empresa objetivo (esto corrige el total/facturas)
            if empresa_objetivo and "Empresa" in df_resumen.columns:
                emp = str(empresa_objetivo).strip()
                before = len(df_resumen)
                df_resumen["Empresa"] = df_resumen["Empresa"].astype(str).str.strip()
                df_resumen = df_resumen[df_resumen["Empresa"] == emp]
                print(f"🏷️ Empresa objetivo: '{emp}' | filas resumen: {before} → {len(df_resumen)}")

            # 3) Filtrar por los lotes equivalentes ya con el resumen filtrado por empresa
            df_lote = self.filtrar_lotes_exactos(df_resumen, posibles_lotes)

            print("📊 Posibles lotes usados para filtrar:")
            print(posibles_lotes)
            print("📋 Contenido CNOMBREPRODUCTO del DataFrame resumen:")
            if "CNOMBREPRODUCTO" in df_resumen.columns:
                print(df_resumen["CNOMBREPRODUCTO"].unique())
            print("🔍 DataFrame resultante después de filtrar_lotes_exactos:")
            if "CNOMBREPRODUCTO" in df_lote.columns:
                print(df_lote["CNOMBREPRODUCTO"].unique())

            if df_lote.empty:
                print("⚠️ No se encontró información del lote seleccionado.")
                return

            # 👉 Estas ya salen SOLO de la empresa objetivo (FACTURAS BASE)
            facturas_base = df_lote["FACTURA"].astype(str).str.strip().dropna().unique().tolist()
            total = float(pd.to_numeric(df_lote["CTOTAL"], errors="coerce").sum())

            print(f"🧾 Facturas base seleccionadas: {facturas_base}")
            print(f"💰 Total acumulado (solo base): {total}")

            # 3.1) === DETECTAR FACTURAS ADICIONALES IGUAL QUE EN generar_pdf_estado_cuenta ===

            # Ubicación / nombre del lote (igual que en el PDF)
            ubicacion = None
            if "CNOMBREPRODUCTO" in df_lote.columns:
                sin_compl = df_lote[~df_lote["CNOMBREPRODUCTO"].astype(str).str.upper().str.contains("COMPL")]
                fila_principal = sin_compl.iloc[0] if not sin_compl.empty else df_lote.iloc[0]
                ubicacion = str(fila_principal["CNOMBREPRODUCTO"])
            print(f"📍 Ubicación detectada para adicionales: {ubicacion}")

            # IDs de documento base (para excluirlos)
            df_lote_tmp = df_lote.copy()
            df_lote_tmp["CTOTAL_NUM"] = pd.to_numeric(df_lote_tmp["CTOTAL"], errors="coerce")
            if "CIDDOCUMENTO" in df_lote_tmp.columns and df_lote_tmp["CIDDOCUMENTO"].notna().any():
                base_doc_ids = set(df_lote_tmp["CIDDOCUMENTO"].dropna().astype(int).unique())
            else:
                base_doc_ids = set()

            print(f"🧾 CIDDOCUMENTO base usados para excluir adicionales: {base_doc_ids}")

            adic_dict, suma_adic = ({}, 0.0)
            if empresa_objetivo and ubicacion:
                try:
                    adic_dict, suma_adic = self.obtener_facturas_adicionales(
                        cliente=cliente,
                        nombre_producto=ubicacion,
                        empresa=empresa_objetivo,
                        excluir_doc_ids=base_doc_ids
                    )
                    print(f"➕ Facturas adicionales detectadas (raw): {adic_dict} (suma={suma_adic})")
                except Exception as e:
                    print(f"❌ Error obteniendo facturas adicionales: {e}")
                    adic_dict, suma_adic = {}, 0.0

            base_set = {s.strip().upper() for s in facturas_base}
            adic_filtrado = {
                etq: tot for etq, tot in adic_dict.items()
                if etq and etq.strip().upper() not in base_set
            }

            facturas_para_pagos = self._ordered_unique(facturas_base + list(adic_filtrado.keys()))

            print(f"🧾 Facturas adicionales filtradas: {list(adic_filtrado.keys())}")
            print(f"📌 Facturas que se usarán para buscar pagos: {facturas_para_pagos}")

            # Actualizamos el total para incluir adicionales (esto afectará la amortización)
            total += float(sum(adic_filtrado.values()))
            print(f"💰 Total global (base + adicionales): {total}")

            # 4) Si aplica, generar programa de pagos (sigue recibiendo SOLO facturas base)
            if resumen_generado and incluir_programa:
                self.generar_tabla_periodo_pagos_cliente(
                    cliente,
                    facturas=facturas_base,
                    empresa_objetivo=empresa_objetivo
                )

            # 5) Recolectar pagos para facturas base + adicionales
            pagos = []

            # IDs globales del cliente (como eliges cliente primero)
            ids_cliente_global = self._resolver_ids_por_nombre_canonico(cliente)
            aliases_cliente = self._obtener_aliases_por_ids(ids_cliente_global) if ids_cliente_global else {self._norm_txt(cliente)}
            aliases_cliente = {self._norm_txt(a) for a in aliases_cliente}  # <<< normaliza

            for empresa in self.empresas:
                if empresa_objetivo and empresa != empresa_objetivo:
                    continue

                tablas = self.tablas_por_empresa[empresa]
                documentos = f'"{tablas["documentos"]}"'

                with sqlite3.connect(self.name_db) as conn:
                    df_doc = pd.read_sql_query(f"""
                        SELECT CFECHA, CFOLIO, CTOTAL, CSERIEDOCUMENTO, CRAZONSOCIAL,
                            CREFERENCIA, CIDDOCUMENTO, CTEXTOEXTRA1, COBSERVACIONES,
                            CIDCLIENTEPROVEEDOR
                        FROM {documentos}
                        WHERE CIDDOCUMENTODE IN (9, 10, 12)
                    """, conn)

                if df_doc.empty:
                    continue

                # Filtro por referencias (facturas del lote + adicionales)
                df_doc["FACTURA_REFERENCIA"] = df_doc["CREFERENCIA"].astype(str).str.strip()
                df_doc = df_doc[df_doc["FACTURA_REFERENCIA"].isin(facturas_para_pagos)]
                if df_doc.empty:
                    continue

                # Filtro por cliente -> preferir ID; si no hay IDs, por cualquiera de sus aliases (ambos nombres)
                if ids_cliente_global:
                    df_doc = df_doc[df_doc["CIDCLIENTEPROVEEDOR"].isin(list(ids_cliente_global))]
                else:
                    df_doc["__NOMBRE__"] = df_doc["CRAZONSOCIAL"].astype(str).str.strip().map(self._norm_txt)
                    df_doc = df_doc[df_doc["__NOMBRE__"].isin(aliases_cliente)]

                if df_doc.empty:
                    continue

                df_doc["EMPRESA"] = empresa
                pagos.append(df_doc)

            if not pagos:
                print("📭 No se encontraron pagos para estas facturas (base + adicionales).")
                messagebox.showerror("Error", f"No se encontraron pagos para estas facturas")
                return

            df_pagos = pd.concat(pagos, ignore_index=True)
            df_pagos["CFECHA"] = pd.to_datetime(df_pagos["CFECHA"], errors="coerce")
            df_pagos = df_pagos.sort_values("CFECHA")
            df_pagos["CFECHA"] = df_pagos["CFECHA"].dt.strftime("%d/%m/%Y")

            # Forma de pago base y final
            df_pagos["Forma de pago base"] = df_pagos.apply(
                lambda row: self.obtener_forma_de_pago(row["CSERIEDOCUMENTO"], row["EMPRESA"]),
                axis=1
            )
            if filtros_forma_pago:
                filtros_norm = {self._norm_txt(fp) for fp in filtros_forma_pago}
                df_pagos["_fpb_norm"] = df_pagos["Forma de pago base"].map(self._norm_txt)
                df_pagos = df_pagos[df_pagos["_fpb_norm"].isin(filtros_norm)].drop(columns=["_fpb_norm"])
                print(f"🔎 Aplicando filtro de forma de pago (normalizado): {filtros_forma_pago}")

            def forma_pago_final(row):
                forma = row["Forma de pago base"]
                if forma != "TRANSFERENCIA":
                    return forma
                extra_raw = row.get("CTEXTOEXTRA1")
                obs = str(row.get("COBSERVACIONES", "")).strip()
                try:
                    extra = int(str(extra_raw).replace(" ", "").strip())
                except (ValueError, TypeError):
                    return "TRANSFERENCIA"
                return self.mapeo.get(extra, "TRANSFERENCIA")

            df_pagos["Forma de pago"] = df_pagos.apply(forma_pago_final, axis=1)

            df_resultado = df_pagos.rename(columns={
                "CFECHA": "Fecha de pago",
                "CFOLIO": "Folio",
                "CTOTAL": "Importe MXN",
                "Forma de pago base": "Forma de pago base"
            })[["Fecha de pago", "Folio", "Importe MXN", "Forma de pago", "Forma de pago base"]].copy()

            df_resultado = df_resultado.reset_index(drop=True)
            df_resultado["FACTURA"] = df_resultado["Folio"].map(
                dict(zip(df_pagos["CFOLIO"], df_pagos["FACTURA_REFERENCIA"]))
            )

            with sqlite3.connect(self.name_db) as conn:
                df_resultado.to_sql("EstadoCuenta_Pagos", conn, index=True, if_exists="replace")
            print("✅ Tabla EstadoCuenta_Pagos creada con éxito (incluye adicionales).")

            # 6) Amortización: usa el total ya ajustado (base + adicionales)
            if incluir_programa:
                try:
                    print("🧮 Intentando generar tabla de amortización...")
                    with sqlite3.connect(self.name_db) as conn:
                        df_pagos_chk = pd.read_sql("SELECT * FROM EstadoCuenta_Pagos", conn)

                    if not df_pagos_chk.empty:
                        abonos_realizados = pd.to_numeric(df_pagos_chk["Importe MXN"], errors="coerce").sum()
                        saldo_sin_interes = total - abonos_realizados
                        print(f"💳 Total factura (base + adicionales): {total}")
                        print(f"💵 Abonos realizados: {abonos_realizados}")
                        print(f"📌 Saldo sin interés: {saldo_sin_interes}")

                        if formula and formula.lower() == "saldo insoluto":
                            with sqlite3.connect(self.name_db) as conn:
                                df_pagares = pd.read_sql("SELECT * FROM EstadoCuenta_PeriodoPagos", conn)
                            if not df_pagares.empty:
                                total_para_amortizacion = float(pd.to_numeric(df_pagares["Saldo"], errors="coerce").iloc[0])
                                enganche_para_amortizacion = 0.0
                            else:
                                print("⚠️ EstadoCuenta_PeriodoPagos está vacío.")
                                total_para_amortizacion = float(total)  # fallback
                                enganche_para_amortizacion = float(enganche or 0)
                        else:
                            total_para_amortizacion = float(total)                  # <— MISMO total que header
                            enganche_para_amortizacion = float(enganche or 0)

                        df_amortizacion = self.generar_tabla_amortizacion_cliente(
                            total_factura=total_para_amortizacion,
                            enganche=enganche_para_amortizacion,
                            tasa_anual=float(tasa_interes or 0),
                            meses=int(meses or 0),
                            fecha_inicio_str=str(fecha_inicio or "01/01/2025"),
                            formula=str(formula or "Saldo total")
                        )

                        with sqlite3.connect(self.name_db) as conn:
                            df_amortizacion.to_sql("EstadoCuenta_Amortizacion", conn, index=False, if_exists="replace")
                        print("📄 Tabla de amortización generada y guardada.")

                except Exception as e:
                    print(f"❌ Error generando tabla de amortización: {e}")

            # 7) Generar PDF (ya le pasas facturas base; el PDF vuelve a detectar adicionales para header)
            pdf_generado = self.generar_pdf_estado_cuenta(
                cliente=cliente,
                lote=lote,
                filtros_forma_pago=filtros_forma_pago,
                incluir_programa=incluir_programa,
                facturas=facturas_base,
                posibles_lotes=posibles_lotes,
                df_lote=df_lote,
                empresa=empresa_objetivo,
                tipo_amortizacion=formula
            )
            return pdf_generado

        except Exception as e:
            print(f"❌ Error generando estado de cuenta completo: {e}")
            messagebox.showerror("Error", f"Error generando estado de cuenta: {e}")
            return False




    def _norm_txt_2(self, s):
        if s is None:
            return ""
        s = str(s).strip().upper()
        # Quita acentos/diacríticos
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        # Colapsa espacios múltiples
        s = re.sub(r"\s+", " ", s)
        return s



    def obtener_forma_de_pago(self, serie, empresa):
        try:
            serie_norm = self._norm_txt_2(serie)
            empresa_norm = self._norm_txt_2(empresa)

            print(f"🔍 Buscando forma de pago | Serie: '{serie_norm}' | Empresa: '{empresa_norm}'")

            # Buscar en la empresa (por nombre normalizado)
            for emp_key, mapa in self.config_pagos.items():
                if self._norm_txt_2(emp_key) == empresa_norm:
                    for keyword, forma_pago in mapa.items():
                        kw_norm = self._norm_txt_2(keyword)
                        if kw_norm in serie_norm:
                            print(f"✅ Match encontrado: {kw_norm} → {forma_pago}")

                            # ✅ Normaliza salida: si es CUENTA DE INCORPORACION (sin acento),
                            # devuélvela siempre como 'CUENTA DE INCORPORACIÓN'
                            if self._norm_txt_2(forma_pago) == self._norm_txt_2("CUENTA DE INCORPORACIÓN"):
                                return "CUENTA DE INCORPORACIÓN"

                            return forma_pago  # resto igual

            # Fallback: buscar en todas las empresas (por si no coincidió el nombre)
            for emp_key, mapa in self.config_pagos.items():
                for keyword, forma_pago in mapa.items():
                    kw_norm = self._norm_txt_2(keyword)
                    if kw_norm in serie_norm:
                        print(f"✅ Match global: {kw_norm} → {forma_pago} (en {emp_key})")

                        if self._norm_txt_2(forma_pago) == self._norm_txt_2("CUENTA DE INCORPORACIÓN"):
                            return "CUENTA DE INCORPORACIÓN"

                        return forma_pago

            print(f"⚠️ Sin coincidencia para serie '{serie}' en empresa '{empresa}'")
            return None

        except Exception as e:
            print(f"❌ Error obteniendo forma de pago para serie '{serie}' en empresa '{empresa}': {e}")
            return None

    
        
    def generar_tabla_estado_cuenta_cliente(self, cliente, posibles_lotes=None):

        print(f"🔍 Buscando lotes asociados al cliente: {cliente}")
        try:
            resultados_lotes = []


            # ===== Helpers de normalización / parsing =====
            def _to_space(s: str) -> str:
                # Convierte TODOS separadores Unicode a espacio, y NBSP/ZWSP explícitos
                return "".join(" " if (unicodedata.category(ch) == "Zs" or ch in ("\u00A0","\u2007","\u202F","\u2009","\u2002","\u2003","\u200B")) else ch for ch in s)

            def _norm(s: str) -> str:
                if not isinstance(s, str): return ""
                s = _to_space(s)
                s = unicodedata.normalize('NFKD', s)
                s = "".join(ch for ch in s if not unicodedata.combining(ch))
                s = re.sub(r"\s+", " ", s.strip())
                return s.upper()

            def _canon_base(s: str) -> str:
                t = _norm(s)
                # Normaliza MZA/LOTE (acepta MANZANA, LOTE2, ceros a la izq)
                t = re.sub(r"(?:\bMZA|MANZANA)\s*0*(\d+)\b", r"MZA \1", t)
                t = re.sub(r"\bLOTE\s*0*(\d+)\b", r"LOTE \1", t)
                t = re.sub(r"\bLOTE(\d+)\b", r"LOTE \1", t)
                t = re.sub(r"\s+", " ", t).strip()
                return t

            def _prefijo_por_nombre(nombre: str) -> str:
                t = _norm(nombre)
                if re.search(r"\bCONSTRUCCI[ÓO]N(ES)?\b", t) or re.search(r"\bCONSTRUCCION(ES)?\b", t):
                    return "C"
                if re.search(r"\bLOTE(S)?\b", t):
                    return "L"
                return "L"

            # Patrones robustos (una sola línea también funciona)
            _pat_priv_token = re.compile(
                r"""PRIV(?:ADA)?\s+(?P<token>[A-Z0-9ÁÉÍÓÚÑ ]+?)\s+(?:MZA|MANZANA)\b""",
                re.IGNORECASE | re.VERBOSE
            )
            _pat_mza = re.compile(r"(?:\bMZA|MANZANA)\s*:?\s*(\d+)", re.IGNORECASE)
            _pat_lote = re.compile(r"\bLOTE\s*:?\s*(\d+)", re.IGNORECASE)
            _pat_simple = re.compile(
                r"""(?P<privada>[A-ZÁÉÍÓÚÑ]+)\s+(?:MZA|MANZANA)\s*:?\s*(?P<mza>\d+)\s+LOTE\s*:?\s*(?P<lote>\d+)""",
                re.IGNORECASE | re.VERBOSE
            )

            _priv_num = {str(int(k)): _norm(v) for k, v in privadas_dict.items()}

            def _extraer_ubicacion(nombre: str):
                """Regresa (privada_norm, mza_int, lote_int) o (None, None, None)."""
                t = _norm(nombre)

                # 1) PRIV/PRIVADA <token> ... (token puede ser nombre o número)
                m = _pat_priv_token.search(t)
                if m:
                    token = _norm(m.group("token"))
                    # MZA y LOTE (últimas ocurrencias en el string para mayor robustez)
                    mza_all = list(_pat_mza.finditer(t))
                    lote_all = list(_pat_lote.finditer(t))
                    if not mza_all or not lote_all:
                        return None, None, None
                    mza = int(mza_all[-1].group(1))
                    lote = int(lote_all[-1].group(1))
                    if token.isdigit():
                        token = _priv_num.get(str(int(token)), token)  # map 5 -> LAS HADAS
                    return _norm(token), mza, lote

                # 2) Rescate: última palabra antes de MZA
                last = None
                for mm in _pat_simple.finditer(t):
                    last = mm
                if last:
                    return _norm(last.group("privada")), int(last.group("mza")), int(last.group("lote"))

                return None, None, None

            # ===== Indexa self.superficies_lotes por prefijo y en formato canónico =====
            # Soporta: "C LAS HADAS MZA 15 LOTE2" o "C LAS HADAS MZA 15 LOTE 2"
            idx_por_prefijo = {"L": {}, "C": {}}
            for k, v in self.superficies_lotes.items():
                k_norm = _norm(k)
                if k_norm.startswith("L "):
                    base = _canon_base(k_norm[2:])
                    try: idx_por_prefijo["L"][base] = float(v)
                    except: idx_por_prefijo["L"][base] = float(str(v).strip())
                elif k_norm.startswith("C "):
                    base = _canon_base(k_norm[2:])
                    try: idx_por_prefijo["C"][base] = float(v)
                    except: idx_por_prefijo["C"][base] = float(str(v).strip())

            # ===== 1) IDs del cliente =====
            ids_canonicos = self._resolver_ids_por_nombre_canonico(cliente)
            ids_texto     = self._resolver_ids_global_por_texto(cliente)
            ids_union     = set(ids_canonicos) | set(ids_texto)

            aliases_cliente = self._obtener_aliases_por_ids(ids_union) if ids_union else {self._norm_txt(cliente)}
            if not aliases_cliente:
                aliases_cliente = { self._norm_txt(cliente) }

            # ===== 2) Recorre empresas =====
            for empresa in self.empresas:
                print(f"\n🏢 Empresa: {empresa}")
                try:
                    tablas = self.tablas_por_empresa[empresa]
                    documentos = f'"{tablas["documentos"]}"'
                    movimientos = f'"{tablas["movimientos"]}"'
                    productos   = f'"{tablas["productos"]}"'

                    # Facturas (ID=4)
                    with sqlite3.connect(self.name_db) as conn:
                        df_facturas = pd.read_sql_query(f"""
                            SELECT CIDDOCUMENTO, CSERIEDOCUMENTO, CFOLIO, CRAZONSOCIAL, CTOTAL, CIDCLIENTEPROVEEDOR
                            FROM {documentos}
                            WHERE CIDDOCUMENTODE = 4
                        """, conn)

                    df_facturas["__NOMBRE__"] = df_facturas["CRAZONSOCIAL"].astype(str).map(self._norm_txt)
                    if ids_union:
                        mask = df_facturas["CIDCLIENTEPROVEEDOR"].isin(list(ids_union)) | df_facturas["__NOMBRE__"].isin(aliases_cliente)
                        df_facturas = df_facturas[mask]
                    else:
                        df_facturas = df_facturas[df_facturas["__NOMBRE__"].isin(aliases_cliente)]

                    print(f"🔎 Facturas encontradas: {len(df_facturas)}")
                    if df_facturas.empty:
                        continue

                    ids_factura = df_facturas["CIDDOCUMENTO"].unique().tolist()
                    placeholders = ",".join(["?"] * len(ids_factura))

                    # Movimientos
                    with sqlite3.connect(self.name_db) as conn:
                        df_mov = pd.read_sql_query(f"""
                            SELECT DISTINCT CIDDOCUMENTO, CIDPRODUCTO
                            FROM {movimientos}
                            WHERE CIDDOCUMENTO IN ({placeholders})
                        """, conn, params=ids_factura)

                    print(f"🧾 Movimientos encontrados: {len(df_mov)}")
                    if df_mov.empty:
                        continue

                    ids_producto = df_mov["CIDPRODUCTO"].unique().tolist()
                    placeholders_prod = ",".join(["?"] * len(ids_producto))

                    # Productos
                    with sqlite3.connect(self.name_db) as conn:
                        df_prod = pd.read_sql_query(f"""
                            SELECT CIDPRODUCTO, CNOMBREPRODUCTO
                            FROM {productos}
                            WHERE CIDPRODUCTO IN ({placeholders_prod})
                        """, conn, params=ids_producto)

                    print("🪪 Productos encontrados:", df_prod["CNOMBREPRODUCTO"].tolist())

                    # Vinculación (NO filtramos por RE_LOTES para no perder casos)
                    df_vinculado = df_mov.merge(df_prod, on="CIDPRODUCTO", how="inner")
                    df_completo  = df_vinculado.merge(df_facturas, on="CIDDOCUMENTO", how="inner")
                    df_completo["FACTURA"] = df_completo["CSERIEDOCUMENTO"].astype(str) + " " + df_completo["CFOLIO"].astype(str)
                    df_completo["Empresa"] = empresa

                    # ===== 3) Prefijo L/C, clave base canónica y lookup =====
                    bases, tipos, superfs = [], [], []
                    for nombre in df_completo["CNOMBREPRODUCTO"].astype(str):
                        pref = _prefijo_por_nombre(nombre)
                        privada, mza, lote = _extraer_ubicacion(nombre)
                        if privada is None:
                            bases.append(None); tipos.append(pref); superfs.append(None); continue
                        base = _canon_base(f"{privada} MZA {mza} LOTE {int(lote)}")
                        val = idx_por_prefijo.get(pref, {}).get(base)

                        # Diagnóstico si no mapeó (muestra claves cercanas y caracteres problemáticos)
                        if val is None:
                            disponibles = list(idx_por_prefijo.get(pref, {}).keys())
                            candidatos = [k for k in disponibles if (_norm(privada) in k) or (f"MZA {mza}" in k)]
                            # imprime caracteres unicode explícitos del nombre original
                            raw = nombre
                            hexes = " ".join(hex(ord(ch)) for ch in raw)
                            print("⚠️ No mapeó:",
                                f"\n  Nombre RAW: {raw}",
                                f"\n  Nombre RAW (hex): {hexes}",
                                f"\n  Prefijo detectado: {pref}",
                                f"\n  Base generada: {base!r}",
                                f"\n  Ejemplos en dict[{pref}] (canónicos): {candidatos[:5]}",
                                sep="")

                        superfs.append(val)
                        bases.append(base); tipos.append(pref)

                    df_completo["PREFIJO_TIPO"]    = tipos
                    df_completo["CLAVE_BASE_NORM"] = bases
                    df_completo["SUPERFICIE"]      = superfs

                    # Resultado parcial + diagnóstico compacto
                    _mask_no = df_completo["SUPERFICIE"].isna() & df_completo["CLAVE_BASE_NORM"].notna()
                    if _mask_no.any():
                        print("🔎 No mapeadas (muestra):")
                        print(df_completo.loc[_mask_no, ["CNOMBREPRODUCTO","PREFIJO_TIPO","CLAVE_BASE_NORM"]]
                            .head(8).to_string(index=False))

                    resultados_lotes.append(
                        df_completo[["Empresa","CRAZONSOCIAL","CNOMBREPRODUCTO","SUPERFICIE","FACTURA","CTOTAL"]]
                        .drop_duplicates()
                    )

                except Exception as e:
                    print(f"❌ Error procesando empresa {empresa}: {e}")
                    continue

            # ===== 4) Consolidado =====
            if resultados_lotes:
                df_resultado = pd.concat(resultados_lotes, ignore_index=True)

                if posibles_lotes:
                    posibles_lotes_limpios = [l.strip().upper() for l in posibles_lotes]
                    df_resultado["CNOMBREPRODUCTO_LIMPIO"] = df_resultado["CNOMBREPRODUCTO"].str.strip().str.upper()
                    df_resultado = df_resultado[df_resultado["CNOMBREPRODUCTO_LIMPIO"].isin(posibles_lotes_limpios)]

                with sqlite3.connect(self.name_db) as conn:
                    df_resultado.to_sql("EstadoCuenta_Resumen", conn, index=False, if_exists="replace")

                print("\n✅ Tabla EstadoCuenta_Resumen guardada correctamente.")
                return True
            else:
                print("\n⚠️ No se encontraron lotes/construcciones asociados al cliente.")
                return False

        except Exception as e:
            print(f"\n❌ Error generando estado de cuenta para cliente {cliente}: {e}")
            return False











    
    
    def generar_tabla_periodo_pagos_cliente(self, cliente, facturas, empresa_objetivo=None):
        """
        Construye EstadoCuenta_PeriodoPagos filtrando pagarés (CIDDOCUMENTODE=15)
        por:
        - CIDCLIENTEPROVEEDOR (IDs del cliente)
        - CREFERENCIA ∈ facturas del lote
        Incluye pagarés aunque CPENDIENTE = 0.
        """
        print(f"📆 Generando periodo de pagos POR FACTURA para cliente: {cliente}")
        try:
            if not facturas:
                print("⚠️ 'facturas' vacío; no se puede filtrar pagarés por referencia.")
                return

            facturas_norm = [self._norm_txt(f) for f in facturas]
            registros = []

            ids_cliente_global = self._resolver_ids_por_nombre_canonico(cliente)
            aliases_cliente = self._obtener_aliases_por_ids(ids_cliente_global) if ids_cliente_global else {self._norm_txt(cliente)}

            for empresa in self.empresas:
                if empresa_objetivo and empresa != empresa_objetivo:
                    continue

                print(f"\n🏢 Empresa: {empresa}")
                try:
                    tablas = self.tablas_por_empresa[empresa]
                    documentos = f'"{tablas["documentos"]}"'

                    with sqlite3.connect(self.name_db) as conn:
                        df_doc = pd.read_sql_query(f"""
                            SELECT CIDDOCUMENTO, CFECHAVENCIMIENTO, CTOTAL, CRAZONSOCIAL,
                                CREFERENCIA, CSERIEDOCUMENTO, CFOLIO, CPENDIENTE,
                                CIDCLIENTEPROVEEDOR
                            FROM {documentos}
                            WHERE CIDDOCUMENTODE = 15
                        """, conn)

                    if df_doc.empty:
                        print("📭 No hay pagarés (tipo 15) en esta empresa.")
                        continue

                    df_doc["__REF__"]    = df_doc["CREFERENCIA"].astype(str).str.strip().str.upper()
                    df_doc["__NOMBRE__"] = df_doc["CRAZONSOCIAL"].astype(str).str.strip().str.upper()

                    # 1) por referencia
                    before = len(df_doc)
                    df_doc = df_doc[df_doc["__REF__"].isin(facturas_norm)]
                    print(f"🎯 Filtrado por referencias: {before} → {len(df_doc)}")
                    if df_doc.empty:
                        continue

                    # 2) por cliente (IDs preferidos; si no, por cualquiera de sus aliases)
                    if ids_cliente_global:
                        df_doc = df_doc[df_doc["CIDCLIENTEPROVEEDOR"].isin(list(ids_cliente_global))]
                    else:
                        df_doc = df_doc[df_doc["__NOMBRE__"].isin(aliases_cliente)]

                    if df_doc.empty:
                        print("⚠️ Tras aplicar filtro por cliente, no hay pagarés en esta empresa.")
                        continue

                    # 3) tabla final (incluye pendientes = 0)
                    df_doc["Fecha de pago"] = pd.to_datetime(df_doc["CFECHAVENCIMIENTO"], errors="coerce", dayfirst=True)
                    df_doc = df_doc.sort_values("Fecha de pago")

                    df_doc["Pago mensual"] = pd.to_numeric(df_doc["CTOTAL"], errors="coerce").fillna(0)
                    df_doc["Pendiente"]    = pd.to_numeric(df_doc["CPENDIENTE"], errors="coerce").fillna(0)
                    df_doc["Acumulado"]    = df_doc["Pago mensual"].cumsum()
                    df_doc["Saldo"]        = df_doc["Pago mensual"][::-1].cumsum()[::-1]

                    df_final_emp = df_doc[["Fecha de pago", "Saldo", "Pago mensual", "Pendiente", "Acumulado"]].copy()
                    df_final_emp.reset_index(drop=True, inplace=True)
                    registros.append(df_final_emp)

                except Exception as e:
                    print(f"❌ Error procesando empresa {empresa} en pagarés: {e}")
                    continue

            if registros:
                df_final = pd.concat(registros, ignore_index=True)
                df_final = df_final.sort_values("Fecha de pago").reset_index(drop=True)
                df_final["Fecha de pago"] = df_final["Fecha de pago"].dt.strftime("%d/%m/%Y")
                df_final["Fecha de pago"] = pd.to_datetime(df_final["Fecha de pago"], errors="coerce", dayfirst=True)

                with sqlite3.connect(self.name_db) as conn:
                    df_final.to_sql("EstadoCuenta_PeriodoPagos", conn, index=False, if_exists="replace")

                print("✅ EstadoCuenta_PeriodoPagos creado/actualizado (pagarés por CREFERENCIA + CIDCLIENTEPROVEEDOR).")
                print(f"📌 Total de filas: {len(df_final)}")
            else:
                print("⚠️ No se encontraron pagarés vinculados a esas facturas para este cliente.")
        except Exception as e:
            print(f"❌ Error generando periodo de pagos (pagarés): {e}")




            
            
    
    

    def generar_tabla_amortizacion_cliente(self, total_factura, enganche, tasa_anual, meses, fecha_inicio_str, formula):
        """
        Genera una tabla de amortización basada en saldo insoluto o saldo total.

        total_factura: float – Monto total facturado
        enganche: float – Pago inicial
        tasa_anual: float – Tasa anual en formato decimal (ej. 0.07 para 7%)
        meses: int – Número de meses del plan
        fecha_inicio_str: str – Fecha inicio en formato 'DD/MM/AAAA'
        formula: str – "Saldo insoluto" o "Saldo total"
        """

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")
        except ValueError:
            raise ValueError("Fecha de inicio inválida. Usa el formato DD/MM/AAAA.")

        saldo_inicial = total_factura - enganche
        if saldo_inicial <= 0:
            raise ValueError("El enganche debe ser menor al total de la factura.")

        tasa_mensual = tasa_anual / 12

        tabla = []
        saldo_restante = saldo_inicial
        acumulado = 0

        if formula.lower() == "saldo insoluto":
            # Calcular pago mensual usando fórmula de anualidad
            pago_mensual = saldo_inicial * (tasa_mensual * (1 + tasa_mensual) ** meses) / ((1 + tasa_mensual) ** meses - 1)     ## 4842.34 / 1.06
        elif formula.lower() == "saldo total":
            # pago_mensual = (saldo_inicial + saldo_inicial * tasa_mensual * meses) / meses
            pago_mensual = saldo_inicial * (tasa_mensual * (1 + tasa_mensual) ** meses) / ((1 + tasa_mensual) ** meses - 1)
        else:
            raise ValueError("Fórmula inválida. Usa 'Saldo insoluto' o 'Saldo total'.")

        for i in range(meses):
            fecha_pago = fecha_inicio + relativedelta(months=i)

            if formula.lower() == "saldo insoluto":
                interes_mensual = saldo_restante * tasa_mensual
            else:
                # interes_mensual = saldo_inicial * tasa_mensual
                interes_mensual = saldo_restante * tasa_mensual

            abono_capital = pago_mensual - interes_mensual
            saldo_mostrado = saldo_restante  # 💡 Guardar antes de restar
            saldo_restante = max(0, saldo_restante - abono_capital)
            acumulado += pago_mensual

            tabla.append({
                "Fecha de pago": fecha_pago.strftime("%d/%m/%Y"),
                "Saldo": round(saldo_mostrado, 2),  # ✅ Ahora correcto
                "Pago mensual": round(pago_mensual, 2),
                "Interés mensual": round(interes_mensual, 2),
                "Abono a capital": round(abono_capital, 2),
                "Acumulado": round(acumulado, 2)
            })


        df_amortizacion = pd.DataFrame(tabla)
        
        with sqlite3.connect(self.name_db) as conn:
            df_amortizacion.to_sql("Tabla_Amortizacion", conn, index=False, if_exists="replace")
        print("✅ Tabla de amortización generada y guardada en la base de datos.")
            
        return df_amortizacion

    


    def generar_pdf_estado_cuenta(self, cliente, lote, filtros_forma_pago=None, incluir_programa=False,
                                facturas=None, empresa=None, posibles_lotes=None, df_lote=None, tipo_amortizacion=None):

            try:
                # --- Lee tablas base ---
                with sqlite3.connect(self.name_db) as conn:
                    df_resumen = pd.read_sql("SELECT * FROM EstadoCuenta_Resumen", conn)
                    df_pagos   = pd.read_sql("SELECT * FROM EstadoCuenta_Pagos", conn)

                # --- Filtra por empresa SOLO en el resumen ---
                if empresa:
                    empresa = str(empresa).strip()
                    if "Empresa" in df_resumen.columns:
                        df_resumen["Empresa"] = df_resumen["Empresa"].astype(str).str.strip()
                        before = len(df_resumen)
                        df_resumen = df_resumen[df_resumen["Empresa"] == empresa]
                        print(f"🏷️ Empresa: '{empresa}' | filas resumen: {before} → {len(df_resumen)}")
                    else:
                        print("⚠️ 'EstadoCuenta_Resumen' no tiene columna Empresa.")

                # --- Construye df_lote SIEMPRE desde el resumen ya filtrado ---
                if facturas is None:
                    posibles_lotes = self.obtener_lotes_equivalentes(cliente, lote)
                    df_lote = self.filtrar_lotes_exactos(df_resumen, posibles_lotes)
                    if df_lote.empty:
                        print("⚠️ No se encontró el lote (ni COMPL) en la empresa seleccionada.")
                        return
                else:
                    df_lote = df_resumen[df_resumen["FACTURA"].isin(facturas)]
                    if df_lote.empty:
                        print("⚠️ No hay info para esas facturas en la empresa seleccionada.")
                        return

                # --- Facturas a mostrar: SIEMPRE desde df_lote ya filtrado ---
                facturas_display = df_lote["FACTURA"].astype(str).str.strip().dropna().tolist()
                facturas_display = self._ordered_unique(facturas_display)   # evitar repetidos (mismo doc en varias filas)
                print(f"🧾 Facturas seleccionadas (base): {facturas_display}")

                # --- Datos base del encabezado ---
                sin_compl = df_lote[~df_lote["CNOMBREPRODUCTO"].str.upper().str.contains("COMPL")]
                fila_principal = sin_compl.iloc[0] if not sin_compl.empty else df_lote.iloc[0]

                ubicacion  = str(fila_principal["CNOMBREPRODUCTO"])
                superficie = float(fila_principal.get("SUPERFICIE", 0) or 0.0)
                
                print(f"📍 Ubicación: {ubicacion} | Superficie: {superficie}")

                # --- Total original: sumar cada factura una sola vez y guardar sus IDs ---
                df_lote_tmp = df_lote.copy()
                df_lote_tmp["CTOTAL_NUM"] = pd.to_numeric(df_lote_tmp["CTOTAL"], errors="coerce")

                if "CIDDOCUMENTO" in df_lote_tmp.columns and df_lote_tmp["CIDDOCUMENTO"].notna().any():
                    base_doc_ids = set(df_lote_tmp["CIDDOCUMENTO"].dropna().astype(int).unique())
                    totales_unicos = (df_lote_tmp.dropna(subset=["CIDDOCUMENTO"])
                                                .drop_duplicates(subset=["CIDDOCUMENTO"]))
                else:
                    base_doc_ids = set()
                    df_lote_tmp["__FACTKEY__"] = df_lote_tmp["FACTURA"].astype(str).str.strip().str.upper()
                    totales_unicos = df_lote_tmp.drop_duplicates(subset=["__FACTKEY__"])

                total_original = float(totales_unicos["CTOTAL_NUM"].sum())
                
                # Obtener el nombre EXACTO del cliente desde un documento base
                cliente_db_exact = None
                try:
                    if base_doc_ids:
                        any_doc = int(next(iter(base_doc_ids)))
                        with sqlite3.connect(self.name_db) as _c:
                            _t_doc = f'"{self.tablas_por_empresa[empresa]["documentos"]}"'
                            _r = pd.read_sql_query(
                                f"SELECT CRAZONSOCIAL FROM {_t_doc} WHERE CIDDOCUMENTO = ? LIMIT 1",
                                _c, params=[any_doc]
                            )
                        if not _r.empty and pd.notna(_r.iloc[0,0]):
                            cliente_db_exact = str(_r.iloc[0,0]).strip()
                except Exception as _e:
                    print("No se pudo leer CRAZONSOCIAL exacto:", _e)

                # === Traer adicionales EXCLUYENDO los documentos base ===
                adic_dict, suma_adic = self.obtener_facturas_adicionales(
                    cliente=cliente,
                    nombre_producto=ubicacion,
                    empresa=empresa,
                    excluir_doc_ids=base_doc_ids    # ← clave para el subquery
                )
                print(f"Las facturas adicionales son: {adic_dict} (suma={suma_adic})")

                base_set = {s.strip().upper() for s in facturas_display}  # base ya dedup

                # Quita adicionales que ya están en la base (misma SERIE FOLIO)
                adic_filtrado = {etq: tot for etq, tot in adic_dict.items() if etq.strip().upper() not in base_set}

                # “Factura” visual
                facturas_full_display = self._ordered_unique(facturas_display + list(adic_filtrado.keys()))
                factura = ", ".join(facturas_full_display)

                # Total GLOBAL
                total_original += sum(adic_filtrado.values())

                # ------------------- << AÑADIR ESTO >> -------------------
                # Total y contexto que debe usar la amortización (el MISMO del header)
                self._amort_total_base = float(total_original)  # total global para amortización
                self._amort_facturas   = facturas_full_display
                self._amort_empresa    = empresa
                # ---------------------------------------------------------

                # ===============================================================

                # --- Pagos: AHORA usamos DIRECTAMENTE la tabla EstadoCuenta_Pagos ---
                # Ya no filtramos por FACTURA, porque esa tabla ya fue filtrada correctamente
                # en generar_estado_cuenta_completo (incluye base + adicionales y cliente correcto).
                df_pagos_lote = df_pagos.copy()

                # Filtros de forma de pago (si aplica) – opcional, por si quieres refinar aún más
                if filtros_forma_pago:
                    if "Forma de pago base" in df_pagos_lote.columns:
                        df_pagos_lote = df_pagos_lote[df_pagos_lote["Forma de pago base"].isin(filtros_forma_pago)]
                    else:
                        print("⚠️ No existe 'Forma de pago base' en PDF. Usando 'Forma de pago'.")
                        df_pagos_lote = df_pagos_lote[df_pagos_lote["Forma de pago"].isin(filtros_forma_pago)]

                # Suma de abonos: ahora es exactamente la misma que verías en la tabla EstadoCuenta_Pagos
                abono = float(pd.to_numeric(df_pagos_lote["Importe MXN"], errors="coerce").sum())

                # --- Subtítulo / total según filtros ---
                filtros_forma_pago = filtros_forma_pago or []
                if set(filtros_forma_pago) != {"EFECTIVO", "TRANSFERENCIA", "CUENTA DE INCORPORACIÓN"} and filtros_forma_pago:
                    # Si estás filtrando por algunas formas específicas, el "total" mostrado es lo abonado
                    total = abono
                    subtitulo = f'ESTADO DE CUENTA CLIENTE - {" / ".join(filtros_forma_pago)}'
                else:
                    # Si no hay filtros (o son todas), el total mostrado es el total global del lote
                    total = total_original
                    subtitulo = "ESTADO DE CUENTA CLIENTE - GLOBAL"

                saldo = float(abono - total)

                # --- Saldo vencido (si existe la tabla) ---
                try:
                    with sqlite3.connect(self.name_db) as conn:
                        df_programa = pd.read_sql("SELECT * FROM EstadoCuenta_PeriodoPagos", conn)
                    df_programa["Fecha de pago"] = pd.to_datetime(df_programa["Fecha de pago"], errors="coerce", dayfirst=True)
                    hoy = pd.Timestamp(datetime.now().date())
                    total_pagares_vencidos = pd.to_numeric(
                        df_programa.loc[df_programa["Fecha de pago"] < hoy, "Pendiente"],
                        errors="coerce"
                    ).sum()
                    saldo_vencido = float(total_pagares_vencidos) if total_pagares_vencidos > 0 else 0.0
                except Exception as e:
                    print(f"❌ Error calculando saldo vencido: {e}")
                    saldo_vencido = 0.0

                # --- Periodo (fechas en string, no Timestamp) ---
                df_pagos_lote["Fecha de pago"] = pd.to_datetime(df_pagos_lote["Fecha de pago"], dayfirst=True, errors="coerce")
                fechas_validas = df_pagos_lote["Fecha de pago"].dropna().sort_values()

                def mes_ano(d):
                    meses = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
                    return f"{meses[d.month]} de {d.year}"

                if not fechas_validas.empty:
                    fi, ff = fechas_validas.iloc[0], fechas_validas.iloc[-1]
                    periodo_reporte = f"Periodo: {mes_ano(fi)}" if fi == ff else f"Periodo: {mes_ano(fi)} - {mes_ano(ff)}"
                else:
                    periodo_reporte = "Periodo: Sin datos"

                # Si luego imprimes la tabla de pagos, vuelve a string la fecha:
                df_pagos_lote["Fecha de pago"] = df_pagos_lote["Fecha de pago"].dt.strftime("%d/%m/%Y")

                # --- PDF ---
                pdf = ReportePDF(titulo_reporte="Estado de Cuenta", periodo_reporte=str(periodo_reporte),
                                empresas=str(empresa or ""))
                pdf.add_page()
                pdf.add_subtitulo(str(subtitulo))
                pdf.agregar_encabezado_estado_cuenta(
                    str(cliente), str(factura), str(ubicacion),
                    float(total), float(abono), float(saldo),
                    float(superficie), float(saldo_vencido), subtitulo
                )

                excel_dict = {}
                
                if "Forma de pago base" in df_pagos_lote.columns:
                    df_pagos_lote = df_pagos_lote.drop(columns=["Forma de pago base"])
                    
                df_pagos_lote = self._normalizar_fecha(df_pagos_lote, "Fecha de pago")
                
                df_pagos_lote_sin_factura = df_pagos_lote.drop(columns=["FACTURA"], errors="ignore")
        
                pdf.agregar_tabla_estado_cuenta(df_pagos_lote_sin_factura)

                try:
                    excel_dict[subtitulo] = df_pagos_lote_sin_factura
                except Exception as e:
                    print("No se encontraron pagos para exportar a Excel:", e)

                if incluir_programa:
                    try:
                        if not df_programa.empty:
                            print("🗓️ Agregando tabla Programa de Pagos al PDF...")
                            
                            if "Pendiente" in df_programa.columns:
                                df_programa = df_programa.drop(columns=["Pendiente"])
                                
                            pdf.add_page()
                            pdf.add_subtitulo('PROGRAMA DE PAGOS')
                            
                            pdf.agregar_tabla_programa_pagos(df_programa)
                            
                            df_programa["Fecha de pago"] = df_programa["Fecha de pago"].dt.strftime("%d/%m/%Y")
                            
                            excel_dict["PROGRAMA DE PAGOS"] = df_programa
                            
                    except Exception as e:
                        print(f"📭 No se pudo agregar Programa de Pagos: {e}")
                        messagebox.showerror("Error", f"No se pudo agregar Programa de Pagos: {e}")

                    # 👇 Amortización
                    try:
                        with sqlite3.connect(self.name_db) as conn:
                            df_amortizacion = pd.read_sql("SELECT * FROM EstadoCuenta_Amortizacion", conn)
                        if not df_amortizacion.empty:
                            print("📄 Agregando tabla de Amortización al PDF...")
                            df_amortizacion["Fecha de pago"] = pd.to_datetime(df_amortizacion["Fecha de pago"], errors="coerce", dayfirst=True)
                    
                            pdf.add_page()
                            tipo_texto = f"{tipo_amortizacion.upper()}" if tipo_amortizacion else ""
                            pdf.add_subtitulo(f"TABLA DE AMORTIZACIÓN - {tipo_texto}")
                            pdf.agregar_tabla_amortizacion(df_amortizacion)
                            
                            excel_dict[f"TABLA DE AMORTIZACIÓN - {tipo_texto}"] = df_amortizacion
                    except Exception as e:
                        print(f"📭 No se pudo agregar Tabla de Amortización: {e}")

                return pdf.guardar_estado_cuenta_pdf(cliente, lote, excel_dict)
            except Exception as e:
                print(f"❌ Error generando PDF de estado de cuenta: {e}")
                return None

        
        
    
    
    def obtener_lotes_por_cliente_y_empresa(self, cliente, empresa):
        """
        En la empresa elegida, busca lotes/construcciones por TODOS los nombres (aliases)
        que comparten el mismo ID del cliente seleccionado en la UI.
        ❗ Aquí NO usamos el ID para filtrar documentos: usamos solo los NOMBRES (aliases).
        """
        try:
            # --- Aliases del mismo ID (normalizados) ---
            ids_cliente = self._resolver_ids_por_nombre_canonico(cliente)
            aliases_norm = self._obtener_aliases_por_ids(ids_cliente, empresa_objetivo=None)
            if not aliases_norm:
                aliases_norm = {self._norm_txt(cliente)}

            print("👥 Aliases usados para búsqueda de lotes:", aliases_norm)

            # --- Tablas ---
            tablas = self.tablas_por_empresa[empresa]
            documentos = f'"{tablas["documentos"]}"'
            movimientos = f'"{tablas["movimientos"]}"'
            productos   = f'"{tablas["productos"]}"'

            # --- Documentos (facturas) y productos ---
            with sqlite3.connect(self.name_db) as conn:
                df_doc = pd.read_sql_query(f"""
                    SELECT CIDDOCUMENTO, CRAZONSOCIAL
                    FROM {documentos}
                    WHERE CIDDOCUMENTODE = 4
                """, conn)

                df_prod = pd.read_sql_query(f"""
                    SELECT CIDPRODUCTO, CNOMBREPRODUCTO
                    FROM {productos}
                """, conn)

            if df_doc.empty:
                return []

            # --- Filtro por aliases (por nombre normalizado) ---
            df_doc["__N__"] = df_doc["CRAZONSOCIAL"].astype(str).map(self._norm_txt)
            df_doc = df_doc[df_doc["__N__"].isin(aliases_norm)]
            if df_doc.empty:
                print("⚠️ Ningún documento coincide con los aliases del cliente.")
                return []

            facturas_ids = df_doc["CIDDOCUMENTO"].unique().tolist()
            placeholders = ",".join(["?"] * len(facturas_ids))

            with sqlite3.connect(self.name_db) as conn:
                df_mov = pd.read_sql_query(f"""
                    SELECT DISTINCT CIDDOCUMENTO, CIDPRODUCTO
                    FROM {movimientos}
                    WHERE CIDDOCUMENTO IN ({placeholders})
                """, conn, params=facturas_ids)

            if df_mov.empty:
                return []

            ids_prod = df_mov["CIDPRODUCTO"].unique().tolist()
            placeholders_prod = ",".join(["?"] * len(ids_prod))
            with sqlite3.connect(self.name_db) as conn:
                df_prod = pd.read_sql_query(f"""
                    SELECT CIDPRODUCTO, CNOMBREPRODUCTO
                    FROM {productos}
                    WHERE CIDPRODUCTO IN ({placeholders_prod})
                """, conn, params=ids_prod)

            if df_prod.empty:
                return []

            # --- Solo lotes/construcciones (evita el warning con grupo no-capturante) ---
            patron = r"(?:LOTE|CONSTRUCCION|ST ANTICIPO)"
            df_lotes = df_prod[df_prod["CNOMBREPRODUCTO"].str.contains(patron, case=False, na=False)].copy()
            if df_lotes.empty:
                print("⚠️ No se encontraron lotes/construcciones para estos nombres.")
                return []

            # --- Quitar " COMPL" visualmente y deduplicar ---
            df_lotes["LoteNormalizado"] = df_lotes["CNOMBREPRODUCTO"].str.replace(" COMPL", "", regex=False).str.strip()
            lotes_unicos = df_lotes["LoteNormalizado"].dropna().unique().tolist()

            print("✅ Lotes encontrados (deduplicados):", lotes_unicos)
            return lotes_unicos

        except Exception as e:
            print(f"❌ Error obteniendo lotes por cliente '{cliente}' y empresa '{empresa}': {e}")
            return []






    def obtener_clientes_unicos_por_id(self, empresa: str) -> list[str]:
        tablas = self.tablas_por_empresa[empresa]
        documentos = f'"{tablas["documentos"]}"'
        with sqlite3.connect(self.name_db) as conn:
            df = pd.read_sql_query(f"""
                SELECT CIDCLIENTEPROVEEDOR, CRAZONSOCIAL
                FROM {documentos}
                WHERE CRAZONSOCIAL IS NOT NULL AND CIDCLIENTEPROVEEDOR IS NOT NULL
            """, conn)

        if df.empty:
            return []

        # Nos quedamos con el nombre "más largo" por cada ID (suele incluir los dos apellidos)
        df["len"] = df["CRAZONSOCIAL"].astype(str).str.len()
        df = df.sort_values(["CIDCLIENTEPROVEEDOR", "len"], ascending=[True, False])
        df = df.drop_duplicates(subset=["CIDCLIENTEPROVEEDOR"], keep="first")
        return sorted(df["CRAZONSOCIAL"].tolist())


  
    
    def obtener_clientes_unicos(self):
        """
        Devuelve la lista global (todas las empresas) de nombres canónicos:
        uno por ID, el MÁS USADO (empate → más largo), considerando 4/9/10/12/15.
        """
        if not hasattr(self, "lista_clientes_ui") or not self.lista_clientes_ui:
            self.construir_catalogo_clientes_global()
        return list(self.lista_clientes_ui)

    
    def obtener_lotes_por_cliente(self, cliente):
        """
        Global (todas las empresas): busca lotes/construcciones por ID del cliente.
        Si no hay ID, intenta por cualquiera de sus aliases (ambos nombres).
        """
        try:
            ids_canonicos = self._resolver_ids_por_nombre_canonico(cliente)
            ids_texto = self._resolver_ids_global_por_texto(cliente)
            ids_union = set(ids_canonicos) | set(ids_texto)

            aliases = self._aliases_de_ids_union(ids_union)
            if not aliases:
                aliases = { self._norm_txt(cliente) }
            resultados  = []

            for empresa in self.empresas:
                tablas = self.tablas_por_empresa[empresa]
                documentos = f'"{tablas["documentos"]}"'
                movimientos = f'"{tablas["movimientos"]}"'
                productos   = f'"{tablas["productos"]}"'

                with sqlite3.connect(self.name_db) as conn:
                    df_fact = pd.read_sql_query(f"""
                        SELECT CIDDOCUMENTO, CIDCLIENTEPROVEEDOR, CRAZONSOCIAL
                        FROM {documentos}
                        WHERE CIDDOCUMENTODE = 4
                    """, conn)

                if ids_union:
                    dff = df_fact[df_fact["CIDCLIENTEPROVEEDOR"].isin(list(ids_union))]
                else:
                    dff = df_fact.copy()
                    dff["__N__"] = dff["CRAZONSOCIAL"].astype(str).str.strip().map(self._norm_txt)
                    dff = dff[dff["__N__"].isin(aliases)]

                if dff.empty: 
                    continue

                ids_doc = dff["CIDDOCUMENTO"].unique().tolist()
                placeholders = ",".join(["?"]*len(ids_doc))

                with sqlite3.connect(self.name_db) as conn:
                    df_mov = pd.read_sql_query(f"""
                        SELECT DISTINCT CIDDOCUMENTO, CIDPRODUCTO
                        FROM {movimientos}
                        WHERE CIDDOCUMENTO IN ({placeholders})
                    """, conn, params=ids_doc)

                if df_mov.empty: 
                    continue

                ids_prod = df_mov["CIDPRODUCTO"].unique().tolist()
                placeholders_prod = ",".join(["?"]*len(ids_prod))
                with sqlite3.connect(self.name_db) as conn:
                    df_prod = pd.read_sql_query(f"""
                        SELECT CIDPRODUCTO, CNOMBREPRODUCTO
                        FROM {productos}
                        WHERE CIDPRODUCTO IN ({placeholders_prod})
                    """, conn, params=ids_prod)

                df_lotes = df_prod[df_prod["CNOMBREPRODUCTO"].str.contains(self.RE_LOTES)]
                resultados.extend(df_lotes["CNOMBREPRODUCTO"].dropna().unique().tolist())

            return sorted(set(resultados))
        except Exception as e:
            print(f"❌ Error obteniendo lotes por cliente '{cliente}': {e}")
            return []
    

    def _norm_txt(self, s: str) -> str:
        if s is None: return ""
        s = str(s).upper().strip()
        s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
        return re.sub(r"\s+", " ", s)

    def _resolver_ids_cliente(self, empresa: str, nombre_usuario: str) -> set[int]:
        """IDs que 'parecen' ese cliente en una empresa (por nombre flexible)."""
        tablas = self.tablas_por_empresa[empresa]
        documentos = f'"{tablas["documentos"]}"'
        target = self._norm_txt(nombre_usuario)
        with sqlite3.connect(self.name_db) as conn:
            df = pd.read_sql_query(f"""
                SELECT DISTINCT CIDCLIENTEPROVEEDOR, CRAZONSOCIAL
                FROM {documentos}
                WHERE CIDCLIENTEPROVEEDOR IS NOT NULL AND CRAZONSOCIAL IS NOT NULL
            """, conn)
        if df.empty: 
            return set()
        df["__N__"] = df["CRAZONSOCIAL"].map(self._norm_txt)
        mask = (df["__N__"] == target) | df["__N__"].str.contains(re.escape(target))
        return set(df.loc[mask, "CIDCLIENTEPROVEEDOR"].astype(int).tolist())

    def construir_catalogo_clientes_global(self):
        """
        Llena:
        - self.lista_clientes_ui : nombres canónicos únicos (el MÁS USADO por ID) para Autocomplete.
        - self.nombre_a_ids      : dict nombre_canónico → set de IDs (CIDCLIENTEPROVEEDOR) en todas las empresas.
        - self.id_a_aliases      : dict id → set de TODOS los nombres observados (para buscar por ambos nombres si hiciera falta).
        Considera Facturas (4), Pagos (9,10,12) y Pagarés (15).
        """
        tipos = (4, 9, 10, 12, 15)
        acumulado = []
        for empresa in self.empresas:
            try:
                tablas = self.tablas_por_empresa[empresa]
                documentos = f'"{tablas["documentos"]}"'
                with sqlite3.connect(self.name_db) as conn:
                    df = pd.read_sql_query(f"""
                        SELECT CIDCLIENTEPROVEEDOR AS ID, CRAZONSOCIAL AS NOMBRE
                        FROM {documentos}
                        WHERE CIDDOCUMENTODE IN ({",".join(map(str, tipos))})
                        AND CIDCLIENTEPROVEEDOR IS NOT NULL
                        AND CRAZONSOCIAL IS NOT NULL
                    """, conn)
                if df.empty:
                    continue
                df["NOMBRE"] = df["NOMBRE"].astype(str).str.strip()
                acumulado.append(df)
            except Exception as e:
                print(f"⚠️ Error leyendo clientes: {e}")
                continue

        if not acumulado:
            self.lista_clientes_ui = []
            self.nombre_a_ids = {}
            self.id_a_aliases = {}
            return

        df_all = pd.concat(acumulado, ignore_index=True)

        # --- Normalización para agrupar bien los aliases ---
        df_all["NORM"] = df_all["NOMBRE"].map(self._norm_txt)

        # Aliases por ID (mantén TODAS las variantes vistas, normalizadas y también crudas)
        # Guardamos upper-normalizado para buscar por nombre sin errores,
        # pero no perdemos la variante cruda (por estética si algún día la quieres).
        id_to_aliases_norm = (
            df_all.groupby("ID")["NORM"]
            .apply(lambda s: set(s.dropna().unique().tolist()))
            .to_dict()
        )

        # Frecuencia (ID, NORM) para escoger canónico “más usado”; empate → más largo
        freq = df_all.groupby(["ID", "NORM"]).size().reset_index(name="CONTEO")
        freq["LEN"] = freq["NORM"].str.len()
        freq = freq.sort_values(["ID", "CONTEO", "LEN"], ascending=[True, False, False])
        canon_por_id = freq.drop_duplicates(subset=["ID"], keep="first")[["ID", "NORM"]]

        # Para lista UI mostramos el canónico pero "bonito".
        # Como no tenemos una versión "bonita" unificada, usamos NORM (upper/si acentos removidos).
        # Si prefieres conservar acentos/caso, aquí podrías mapear de vuelta a la variante más larga original.
        self.lista_clientes_ui = sorted(canon_por_id["NORM"].unique().tolist())

        # nombre (canónico) -> IDs
        self.nombre_a_ids = (
            canon_por_id.groupby("NORM")["ID"]
            .apply(lambda s: set(map(int, s.tolist())))
            .to_dict()
        )

        # id -> set de aliases (upper normalizados)
        self.id_a_aliases = {int(k): set(v) for k, v in id_to_aliases_norm.items()}

        print(f"👤 Catálogo global listo: {len(self.lista_clientes_ui)} clientes únicos.")


    def _resolver_ids_por_nombre_canonico(self, nombre_ui: str) -> set[int]:
        if not hasattr(self, "nombre_a_ids") or not self.nombre_a_ids:
            self.construir_catalogo_clientes_global()
        clave = self._norm_txt(nombre_ui or "")
        return set(self.nombre_a_ids.get(clave, set()))


    def _obtener_aliases_por_ids(self, ids: set[int], empresa_objetivo: str | None = None) -> set[str]:
        """
        Devuelve TODOS los nombres (CRAZONSOCIAL) observados para esos IDs.
        Siempre une:
        - lo del caché (self.id_a_aliases)
        - + una consulta a BD (para no perder variantes que no quedaron en caché)
        Retorna SIEMPRE nombres normalizados con _norm_txt (upper, sin acentos/espacios extra).
        """
        aliases = set()

        # 1) Caché global
        if hasattr(self, "id_a_aliases") and self.id_a_aliases:
            for i in ids:
                for a in self.id_a_aliases.get(int(i), set()):
                    aliases.add(self._norm_txt(a))

        # 2) Consulta directa por seguridad
        if ids:
            placeholders = ",".join(["?"] * len(ids))
            for empresa in self.empresas:
                if empresa_objetivo and empresa != empresa_objetivo:
                    continue
                try:
                    tablas = self.tablas_por_empresa[empresa]
                    documentos = f'"{tablas["documentos"]}"'
                    with sqlite3.connect(self.name_db) as conn:
                        df = pd.read_sql_query(f"""
                            SELECT DISTINCT CRAZONSOCIAL
                            FROM {documentos}
                            WHERE CIDCLIENTEPROVEEDOR IN ({placeholders})
                            AND CRAZONSOCIAL IS NOT NULL
                        """, conn, params=list(ids))
                    if not df.empty:
                        aliases |= set(df["CRAZONSOCIAL"].astype(str).map(self._norm_txt).tolist())
                except Exception as e:
                    print(f"⚠️ Alias fallback error en {empresa}: {e}")

        return aliases
    
    def obtener_facturas_adicionales(self, cliente: str, nombre_producto: str, empresa: str,
                                    excluir_doc_ids=None, devolver_ids=False,
                                    cliente_db_exact: str | None = None):
        excluir_doc_ids = set(excluir_doc_ids or [])

        def _ret(dic, total, ids=set()):
            return (dic, float(total or 0.0), set(ids)) if devolver_ids else (dic, float(total or 0.0))

        if empresa not in self.tablas_por_empresa:
            return _ret({}, 0.0, set())

        tablas = self.tablas_por_empresa[empresa]
        t_prod = f'"{tablas["productos"]}"'
        t_mov  = f'"{tablas["movimientos"]}"'
        t_doc  = f'"{tablas["documentos"]}"'

        with sqlite3.connect(self.name_db) as conn:
            q_cod = f'SELECT CCODIGOPRODUCTO FROM {t_prod} WHERE UPPER(TRIM(CNOMBREPRODUCTO))=UPPER(TRIM(?)) LIMIT 1'
            r = pd.read_sql_query(q_cod, conn, params=[nombre_producto])
            if r.empty:
                return _ret({}, 0.0, set())
            codigo_base = (r.iloc[0, 0] or "").strip().upper()

            def _norm(s):
                return ' '.join(''.join(ch for ch in unicodedata.normalize('NFD', str(s).upper().strip())
                                        if unicodedata.category(ch) != 'Mn').split())
                
            def _norm_txt(s: str) -> str:
                if s is None: return ""
                s = str(s).strip().upper()
                s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
                return ' '.join(s.split())

            es_construccion = ("CONSTRUCCION CASA RESID" in _norm(nombre_producto)
                            or "CONSTRUCCION LOTE HADAS" in _norm(nombre_producto))

            if es_construccion:
                q_adic = f"""
                    SELECT DISTINCT d.CIDDOCUMENTO, d.CSERIEDOCUMENTO, d.CFOLIO, d.CTOTAL, d.CRAZONSOCIAL
                    FROM {t_mov} m
                    JOIN {t_prod} p ON p.CIDPRODUCTO = m.CIDPRODUCTO
                    JOIN {t_doc}  d ON d.CIDDOCUMENTO = m.CIDDOCUMENTO
                    WHERE (
                            (UPPER(p.CCODIGOPRODUCTO) LIKE ? AND UPPER(p.CCODIGOPRODUCTO) <> ?)
                        OR  UPPER(p.CCODIGOPRODUCTO) = 'S1560'
                        )
                    AND d.CIDDOCUMENTODE = 4
                """
                params = [codigo_base + "%", codigo_base]
            else:
                q_adic = f"""
                    SELECT DISTINCT d.CIDDOCUMENTO, d.CSERIEDOCUMENTO, d.CFOLIO, d.CTOTAL, d.CRAZONSOCIAL
                    FROM {t_mov} m
                    JOIN {t_prod} p ON p.CIDPRODUCTO = m.CIDPRODUCTO
                    JOIN {t_doc}  d ON d.CIDDOCUMENTO = m.CIDDOCUMENTO
                    WHERE (UPPER(p.CCODIGOPRODUCTO) LIKE ? AND UPPER(p.CCODIGOPRODUCTO) <> ?)
                    AND d.CIDDOCUMENTODE = 4
                """
                params = [codigo_base + "%", codigo_base]

            df_adic = pd.read_sql_query(q_adic, conn, params=params)
            
        if not df_adic.empty:
            df_adic["CRAZONSOCIAL_N"] = df_adic["CRAZONSOCIAL"].apply(_norm_txt)
            cliente_n = _norm_txt(cliente)
            df_adic = df_adic[df_adic["CRAZONSOCIAL_N"] == cliente_n]
            
            
            print("Código base:", codigo_base)
            print("Params SQL:", params)
            print(df_adic)

        if df_adic.empty:
            return _ret({}, 0.0, set())
        

        if excluir_doc_ids:
            df_adic = df_adic[~df_adic["CIDDOCUMENTO"].astype(int).isin(excluir_doc_ids)]
            if df_adic.empty:
                return _ret({}, 0.0, set())

        df_adic = df_adic.drop_duplicates(subset=["CIDDOCUMENTO"])

        adic_dict, doc_ids = {}, set()
        for _, row in df_adic.iterrows():
            etiqueta = f'{str(row["CSERIEDOCUMENTO"]).strip()} {str(row["CFOLIO"]).strip()}'
            adic_dict[etiqueta] = float(row["CTOTAL"] or 0.0)
            doc_ids.add(int(row["CIDDOCUMENTO"]))

        return _ret(adic_dict, sum(adic_dict.values()), doc_ids)




    

    def _normalize_txt(self, s: str) -> str:
        if s is None:
            return ""
        s = str(s).strip().upper()
        # quitar acentos
        s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
        # colapsar espacios
        s = ' '.join(s.split())
        return s

    def obtener_clientes_de_todas_las_empresas(self, name_db, tablas_por_empresa) -> list:
        clientes_raw = set()
        with sqlite3.connect(name_db) as conn:
            for emp, tablas in tablas_por_empresa.items():
                t_doc = f'"{tablas["documentos"]}"'
                try:
                    df = pd.read_sql_query(
                        f"""
                        SELECT DISTINCT CRAZONSOCIAL
                        FROM {t_doc}
                        WHERE CRAZONSOCIAL IS NOT NULL AND CRAZONSOCIAL <> ''
                        AND CIDDOCUMENTODE = 4
                        """,
                        conn
                    )
                    for c in df["CRAZONSOCIAL"].dropna().tolist():
                        clientes_raw.add(str(c).strip())  # <-- guarda el “bonito”
                except Exception as e:
                    print(f"⚠️ No pude leer clientes de {emp}: {e}")
        return sorted(clientes_raw)
    
    def _ordered_unique(self, items):
        seen = set(); out = []
        for x in items:
            k = str(x).strip().upper()
            if k in seen:
                continue
            seen.add(k)
            out.append(str(x).strip())
        return out

    def _sql_unaccent(self, token: str) -> str:
        """
        Devuelve una expresión SQL que normaliza MAYÚSCULAS+TRIM y elimina acentos
        (ÁÀÂÃÄ→A, ÉÈÊË→E, ... Ü→U, Ñ→N) para comparar insensible a acentos/ñ.
        `token` puede ser un nombre de columna (ej. "d.CRAZONSOCIAL") o "?".
        """
        expr = f"UPPER(TRIM({token}))"
        for a, b in [
            ('Á','A'),('À','A'),('Â','A'),('Ã','A'),('Ä','A'),
            ('É','E'),('È','E'),('Ê','E'),('Ë','E'),
            ('Í','I'),('Ì','I'),('Î','I'),('Ï','I'),
            ('Ó','O'),('Ò','O'),('Ô','O'),('Õ','O'),('Ö','O'),
            ('Ú','U'),('Ù','U'),('Û','U'),('Ü','U'),
            ('Ñ','N'),
            # también minúsculas por si el COLLATION no sube bien algunas
            ('á','A'),('à','A'),('â','A'),('ã','A'),('ä','A'),
            ('é','E'),('è','E'),('ê','E'),('ë','E'),
            ('í','I'),('ì','I'),('î','I'),('ï','I'),
            ('ó','O'),('ò','O'),('ô','O'),('õ','O'),('ö','O'),
            ('ú','U'),('ù','U'),('û','U'),('ü','U'),
            ('ñ','N'),
        ]:
            expr = f"REPLACE({expr}, '{a}', '{b}')"
        return expr


    # GUARDADO DE EXCEL EN FORMATO REPORTE #

    # --- Conversión de unidades para Excel ---
    def mm_to_px(self, mm, dpi=96):        # Excel/openpyxl trabaja bien con ~96 dpi
        return int((mm / 25.4) * dpi)

    def mm_to_points(self, mm):            # 1 point = 1/72 in
        return (mm / 25.4) * 72

    def generar_excel_estado_cuenta_like_pdf(
        self,
        cliente: str,
        lote: str,
        empresa: str,
        periodo_texto: str,
        departamento: str = "Administración",
        logo_path: str = None,
        titulo_reporte: str = "Estado de cuenta",
        subtitulo_estado: str = "ESTADO DE CUENTA CLIENTE - GLOBAL"
    ):
        """
        Genera un Excel con el layout del PDF del Estado de Cuenta en una sola hoja.
        - Encabezado (logo + datos de reporte)
        - Subtítulo del estado
        - Bloque de datos del cliente (Cliente, Factura, Ubicación, Superficie, Total, Abono, Saldo, Saldo vencido)
        - Tabla de pagos (EstadoCuenta_Resumen)  [OBLIGATORIA]
        - Programa de pagos (EstadoCuenta_PeriodoPagos)  [SI EXISTE]
        - Tabla de amortización (EstadoCuenta_Amortizacion)  [SI EXISTE]
        """
        try:
            # ========= Carga de datos =========
            with sqlite3.connect(self.name_db) as conn:
                df_resumen = pd.read_sql_query("SELECT * FROM EstadoCuenta_Resumen", conn)
                # Estas pueden no existir según el caso
                try:
                    df_prog = pd.read_sql_query("SELECT * FROM EstadoCuenta_PeriodoPagos", conn)
                except Exception:
                    df_prog = pd.DataFrame()
                try:
                    df_amort = pd.read_sql_query("SELECT * FROM EstadoCuenta_Amortizacion", conn)
                except Exception:
                    df_amort = pd.DataFrame()

            # Filtra el cliente (y si tienes empresa en la tabla, puedes filtrar también)
            df_cli = df_resumen.copy()
            if "CLIENTE" in df_cli.columns:
                df_cli = df_cli[df_cli["CLIENTE"].astype(str).str.upper() == cliente.upper()]

            if df_cli.empty:
                messagebox.showwarning("Sin datos", "No se encontraron registros del cliente para exportar.")
                return

            # Datos del bloque (ajusta nombres si difieren en tu tabla)
            # Si ya calculas estos totales en otro lado, puedes leerlos de ahí:
            total = float(df_cli["TOTAL"].sum()) if "TOTAL" in df_cli.columns else 0.0
            abono = float(df_cli["ABONO"].sum()) if "ABONO" in df_cli.columns else 0.0
            saldo = float(df_cli["SALDO"].sum()) if "SALDO" in df_cli.columns else total - abono
            saldo_vencido = float(df_cli["SALDO_VENCIDO"].sum()) if "SALDO_VENCIDO" in df_cli.columns else 0.0

            # Si en tu resumen tienes una sola factura/ubicación/superficie consolidada, intenta leerla:
            factura_txt = ""
            ubicacion_txt = ""
            superficie_val = None
            for col_name in ["FACTURA", "FACTURA_TXT", "Factura"]:
                if col_name in df_cli.columns:
                    factura_txt = str(df_cli.iloc[0][col_name])
                    break
            for col_name in ["UBICACION", "UBICACIÓN", "Ubicacion", "Ubicación"]:
                if col_name in df_cli.columns:
                    ubicacion_txt = str(df_cli.iloc[0][col_name])
                    break
            for col_name in ["SUPERFICIE", "Superficie"]:
                if col_name in df_cli.columns:
                    try:
                        superficie_val = float(df_cli.iloc[0][col_name])
                    except Exception:
                        superficie_val = None
                    break

            # ========= Libro y hoja =========
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Estado de Cuenta"

            # Config impresión como el PDF
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # alto variable
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.oddFooter.center.text = "Página &P"

            # Columnas anchas tipo PDF
            for c in range(1, 11):
                ws.column_dimensions[get_column_letter(c)].width = 18

            # ========= Estilos =========
            F_B = Font(bold=True)
            F_TITLE = Font(bold=True, size=14)
            F_SUB = Font(bold=True, size=12)
            A_C = Alignment(horizontal="center", vertical="center")
            A_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
            A_R = Alignment(horizontal="right", vertical="center")
            F_RED = Font(color="FF0000", bold=True)
            F_GRAY = PatternFill("solid", fgColor="EEEEEE")
            BORDER = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            fila = 1

            # === Parámetros “como en el PDF” ===
            logo_path = self.ruta_recurso(LOGO_FILENAME)   # igual que en tu PDF
            logo_width_mm = 40                        # logo_width
            texto_x_mm = 60                           # texto_x (distancia desde el borde izquierdo a texto)
            num_lineas = 7
            alto_linea_pt = 6                         # alto_linea (en puntos, tal cual usas en PDF)
            header_height_mm = 40

            # === Preparación de rejilla para que se parezca al PDF ===
            # 1) Altura del bloque de encabezado
            alto_header_pts = self.mm_to_points(header_height_mm)
            # Usaremos filas 1..num_lineas para el bloque
            for r in range(1, num_lineas + 1):
                ws.row_dimensions[r].height = alto_linea_pt  # mismo alto por línea, como en el PDF
            # Si quieres respetar estrictamente header_height, reparte el “sobrante”:
            sobrante = max(alto_header_pts - (num_lineas * alto_linea_pt), 0)
            if sobrante > 0:
                ws.row_dimensions[num_lineas].height += sobrante  # compensa en la última línea

            # 2) Reserva área de logo y de texto “como texto_x”
            #    Aproximamos “texto_x_mm” a un ancho de columnas.
            #    Truco: usa columnas A..C para el logo y arranca texto en D.
            #    Ajusta los anchos para que ~texto_x_mm.
            px_total = self.mm_to_px(texto_x_mm)  # ancho deseado hasta el texto
            # Excel: 1 “unidad de ancho” de columna ≈ 7 pix aprox (depende de fuente). Usamos 7 como base.
            def px_to_colw(px): return px / 7.0
            ancho_logo_bloque_px = max(px_total, self.mm_to_px(logo_width_mm) + 12)  # deja margen
            colw_A = px_to_colw(ancho_logo_bloque_px * 0.55)
            colw_B = px_to_colw(ancho_logo_bloque_px * 0.30)
            colw_C = px_to_colw(ancho_logo_bloque_px * 0.15)

            ws.column_dimensions["A"].width = colw_A
            ws.column_dimensions["B"].width = colw_B
            ws.column_dimensions["C"].width = colw_C
            # El texto comenzará en la columna D (≈ texto_x)

            # 3) Inserta logo escalado a logo_width_mm
            try:
                img = XLImage(logo_path)
                # Escala manteniendo proporción según ancho deseado
                ancho_px = self.mm_to_px(logo_width_mm)
                ratio = ancho_px / img.width
                img.width = ancho_px
                img.height = int(img.height * ratio)
                ws.add_image(img, "A1")  # esquina superior izquierda
            except Exception:
                pass

            # 4) Bloque de texto a la derecha (desde columna D), una línea por renglón
            #    Misma estructura que tu PDF:
            fila_txt = 1
            def enc_linea(etq, val, fila):
                ws.merge_cells(f"D{fila}:E{fila}")
                ws.merge_cells(f"F{fila}:I{fila}")
                ws[f"D{fila}"] = etq; ws[f"D{fila}"].alignment = A_L; ws[f"D{fila}"].font = F_B
                ws[f"F{fila}"] = val; ws[f"F{fila}"].alignment = A_L

            ws.merge_cells(f"D{fila_txt}:I{fila_txt}")
            ws[f"D{fila_txt}"] = "Grupo Las Hadas"
            ws[f"D{fila_txt}"].font = F_B
            ws[f"D{fila_txt}"].alignment = A_L
            fila_txt += 1

            enc_linea("Reporte:", "Estado de cuenta", fila_txt); fila_txt += 1
            enc_linea("Empresas:", empresa, fila_txt); fila_txt += 1
            enc_linea("Periodo:", periodo_texto, fila_txt); fila_txt += 1
            enc_linea("Fecha de generación:", datetime.now().strftime("%d de %B de %Y"), fila_txt); fila_txt += 1
            enc_linea("Departamento:", "Administración", fila_txt); fila_txt += 1
            # (Si te sobra 1 línea para completar num_lineas, puedes dejarla vacía)

            # ========= Subtítulo del estado =========
            ws.merge_cells(f"A{fila}:I{fila}")
            ws[f"A{fila}"] = subtitulo_estado
            ws[f"A{fila}"].font = F_SUB
            ws[f"A{fila}"].alignment = A_L
            fila += 2

            # ========= Bloque de datos (dos columnas tipo PDF) =========
            def _kv(r, c_label, label, value, value_is_money=False, value_color_red_if_negative=False):
                # r = fila base, c_label = columna letra para etiqueta ("A" o "E")
                c_val = chr(ord(c_label) + 1)
                ws[f"{c_label}{r}"] = label
                ws[f"{c_label}{r}"].font = F_B
                ws[f"{c_label}{r}"].alignment = A_L

                cell = ws[f"{c_val}{r}"]
                cell.value = value
                cell.alignment = A_L

                if value_is_money and isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
                if value_color_red_if_negative and isinstance(value, (int, float)) and value < 0:
                    cell.font = F_RED

            # Columna izquierda
            base = fila
            _kv(base + 0, "A", "Cliente", cliente)
            _kv(base + 1, "A", "Factura", factura_txt)  # si es larga, Excel envuelve
            _kv(base + 2, "A", "Ubicación", ubicacion_txt)
            sup_txt = f"{superficie_val:.2f} m²" if superficie_val is not None else "0.00 m²"
            _kv(base + 3, "A", "Superficie", sup_txt)

            # Columna derecha (con importes)
            _kv(base + 0, "E", "Saldo", saldo, value_is_money=True, value_color_red_if_negative=True)
            _kv(base + 1, "E", "Saldo vencido", saldo_vencido, value_is_money=True, value_color_red_if_negative=True)
            _kv(base + 2, "E", "Total", total, value_is_money=True)
            _kv(base + 3, "E", "Abono", abono, value_is_money=True)

            fila = base + 5

            # ========= Helper para secciones de tabla =========
            def _titulo_seccion(texto):
                nonlocal fila
                ws.merge_cells(f"A{fila}:I{fila}")
                ws[f"A{fila}"] = texto
                ws[f"A{fila}"].font = F_SUB
                ws[f"A{fila}"].alignment = A_L
                fila += 1

            def _render_tabla(df: pd.DataFrame, col_anchos=None, money_cols=None, center_cols=None):
                """Pinta encabezados con gris y bordes; aplica formato moneda a 'money_cols'."""
                nonlocal fila
                if df is None or df.empty:
                    return

                # Encabezados
                for j, col in enumerate(df.columns, start=1):
                    cell = ws.cell(row=fila, column=j, value=str(col))
                    cell.font = F_B
                    cell.alignment = A_C
                    cell.fill = F_GRAY
                    cell.border = BORDER
                fila += 1

                # Filas
                for _, row in df.iterrows():
                    for j, col in enumerate(df.columns, start=1):
                        val = row[col]
                        cell = ws.cell(row=fila, column=j, value=val)
                        cell.alignment = A_L
                        # Moneda
                        if money_cols and col in money_cols and pd.notna(val):
                            try:
                                cell.value = float(val)
                                cell.number_format = '$#,##0.00'
                                cell.alignment = A_R
                            except Exception:
                                pass
                        # Centrado si se pide
                        if center_cols and col in center_cols:
                            cell.alignment = A_C
                        cell.border = BORDER
                    fila += 1

                fila += 1  # espacio

                # Anchos específicos opcionales
                if col_anchos:
                    for idx, w in enumerate(col_anchos, start=1):
                        ws.column_dimensions[get_column_letter(idx)].width = w

            # ========= TABLA 1: Pagos (Resumen) =========
            _titulo_seccion("")  # línea fina como en el PDF (opcional dejar vacío)
            _titulo_seccion("")  # doble línea
            _titulo_seccion("")

            _titulo_seccion("ESTADO DE CUENTA – PAGOS")
            # Ordena/selecciona columnas como en tu PDF (#, Fecha pago, Folio, Importe MXN, Forma de pago)
            # Ajusta nombres de columnas reales:
            cols_pdf = []
            for name in ["#", "Fecha pago", "Folio", "Importe MXN", "Forma de pago"]:
                # Si tus nombres en df_cli son distintos, mapea aquí:
                if name in df_cli.columns:
                    cols_pdf.append(name)
            # Si no existen con esos nombres, intenta mapear:
            if not cols_pdf:
                mapeo = {
                    "#": "#",
                    "Fecha pago": "FECHA_PAGO" if "FECHA_PAGO" in df_cli.columns else "FECHA",
                    "Folio": "FOLIO" if "FOLIO" in df_cli.columns else "CFOLIO" if "CFOLIO" in df_cli.columns else None,
                    "Importe MXN": "IMPORTE" if "IMPORTE" in df_cli.columns else "ABONO",
                    "Forma de pago": "FORMA_PAGO" if "FORMA_PAGO" in df_cli.columns else "FORMAPAGO" if "FORMAPAGO" in df_cli.columns else None,
                }
                cols_pdf = [c for c in ["#", "Fecha pago", "Folio", "Importe MXN", "Forma de pago"] if mapeo.get(c)]
                df_pagos = pd.DataFrame()
                for etiqueta in cols_pdf:
                    df_pagos[etiqueta] = df_cli[mapeo[etiqueta]]
            else:
                df_pagos = df_cli[cols_pdf].copy()

            # Número correlativo si no existe "#"
            if "#" not in df_pagos.columns:
                df_pagos.insert(0, "#", range(1, len(df_pagos) + 1))

            _render_tabla(
                df_pagos,
                col_anchos=[6, 16, 12, 18, 28],
                money_cols={"Importe MXN"},
                center_cols={"#", "Folio"}
            )

            # ========= TABLA 2: Programa de pagos (si existe) =========
            if not df_prog.empty:
                _titulo_seccion("PROGRAMA DE PAGOS")
                # En tu PDF: # | Fecha de pago | Saldo | Pago mensual | Acumulado
                # Acomoda si tus columnas tienen otros nombres
                posibles = ["#", "Fecha de pago", "Saldo", "Pago mensual", "Acumulado"]
                mapeo = {}
                for p in posibles:
                    if p in df_prog.columns:
                        mapeo[p] = p
                # fallback por nombres alternos
                if "Fecha de pago" not in mapeo:
                    for alt in ["FECHA_PAGO", "FECHA", "Fecha"]:
                        if alt in df_prog.columns:
                            mapeo["Fecha de pago"] = alt; break
                if "Saldo" not in mapeo:
                    for alt in ["SALDO", "Saldo"]:
                        if alt in df_prog.columns:
                            mapeo["Saldo"] = alt; break
                if "Pago mensual" not in mapeo:
                    for alt in ["PAGO_MENSUAL", "PAGO", "PAGO_MENSUAL_MXN"]:
                        if alt in df_prog.columns:
                            mapeo["Pago mensual"] = alt; break
                if "Acumulado" not in mapeo:
                    for alt in ["ACUMULADO", "Acumulado"]:
                        if alt in df_prog.columns:
                            mapeo["Acumulado"] = alt; break

                orden = [c for c in posibles if c in mapeo]
                dfp = pd.DataFrame()
                for c in orden:
                    dfp[c] = df_prog[mapeo[c]]

                if "#" not in dfp.columns:
                    dfp.insert(0, "#", range(1, len(dfp) + 1))

                _render_tabla(
                    dfp,
                    col_anchos=[6, 18, 18, 18, 18],
                    money_cols={"Saldo", "Pago mensual", "Acumulado"},
                    center_cols={"#"}
                )

            # ========= TABLA 3: Amortización (si existe) =========
            if not df_amort.empty:
                # Título con variante (“SALDO TOTAL” / “SALDO INSOLUTO”) si ya lo calculas
                titulo_amort = "TABLA DE AMORTIZACIÓN"
                for col in df_amort.columns:
                    if "SALDO TOTAL" in str(col).upper():
                        titulo_amort = "TABLA DE AMORTIZACIÓN - SALDO TOTAL"
                        break
                    if "SALDO INSOLUTO" in str(col).upper():
                        titulo_amort = "TABLA DE AMORTIZACIÓN - SALDO INSOLUTO"
                        break
                _titulo_seccion(titulo_amort)

                # En tu PDF: # | Fecha de pago | Saldo | Pago mensual | Interés mensual | Abono a capital | Acumulado
                posibles = ["#", "Fecha de pago", "Saldo", "Pago mensual", "Interés mensual", "Abono a capital", "Acumulado"]
                mapeo = {}
                for p in posibles:
                    if p in df_amort.columns:
                        mapeo[p] = p
                # Fallbacks típicos
                alt_map = {
                    "Fecha de pago": ["FECHA_PAGO", "FECHA"],
                    "Saldo": ["SALDO"],
                    "Pago mensual": ["PAGO_MENSUAL", "PAGO"],
                    "Interés mensual": ["INTERES_MENSUAL", "INTERÉS", "INTERES"],
                    "Abono a capital": ["ABONO_CAPITAL", "ABONO A CAPITAL"],
                    "Acumulado": ["ACUMULADO"],
                }
                for clave, alts in alt_map.items():
                    if clave not in mapeo:
                        for alt in alts:
                            if alt in df_amort.columns:
                                mapeo[clave] = alt
                                break

                orden = [c for c in posibles if c in mapeo]
                dfa = pd.DataFrame()
                for c in orden:
                    dfa[c] = df_amort[mapeo[c]]

                if "#" not in dfa.columns:
                    dfa.insert(0, "#", range(1, len(dfa) + 1))

                _render_tabla(
                    dfa,
                    col_anchos=[6, 18, 18, 18, 18, 18, 18],
                    money_cols={"Saldo", "Pago mensual", "Interés mensual", "Abono a capital", "Acumulado"},
                    center_cols={"#"}
                )

            # ========= Guardar =========
            fecha_tag = datetime.now().strftime("%Y-%m-%d")
            nombre = f"Estado de Cuenta - {cliente} - {lote} - {fecha_tag}.xlsx"
            ruta = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel (*.xlsx)", "*.xlsx")],
                initialfile=nombre,
                title="Guardar Estado de Cuenta (Excel)"
            )
            if ruta:
                wb.save(ruta)
                messagebox.showinfo("Éxito", f"Excel generado:\n{ruta}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar el Excel:\n{e}")
